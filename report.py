import logging
from typing import List, Dict, Any
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from db import get_session

logger = logging.getLogger(__name__)


def run_query(query: str, params: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Run a SQL query and return rows as dicts."""
    logger.info("Executing query: %s", query)
    stmt = text(query)
    with get_session() as session:
        result = session.execute(stmt, params)
        return [dict(row) for row in result]


def create_excel(data: List[Dict[str, Any]], path: str) -> None:
    """Generate an Excel report with simple header styling."""
    if not data:
        logger.warning("No data provided for Excel report")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = list(data[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in data:
        ws.append([row[h] for h in headers])

    # Auto-width columns
    for column in range(1, ws.max_column + 1):
        length = max(len(str(cell.value)) for cell in ws[get_column_letter(column)]) + 2
        ws.column_dimensions[get_column_letter(column)].width = length

    wb.save(path)
    logger.info("Saved report to %s", path)


def generate_report(query_name: str, sql: str, email: str) -> str:
    """Run query for an email and create an Excel report."""
    data = run_query(sql, {"email": email})
    report_path = f"{query_name}_{email.replace('@', '_')}.xlsx"
    create_excel(data, report_path)
    return report_path
